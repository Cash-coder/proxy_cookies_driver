from openpyxl import load_workbook
import undetected_chromedriver.v2 as uc
from selenium.webdriver.common.by import By

FILE = r'C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\login and launcher\id.xlsx'

wb = load_workbook(filename = FILE)
ws1 = wb['Sheet1']
ws2 = wb['Sheet2']

columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U']
for row in ws1.iter_rows(values_only=True):
    name = row[0]
    id = row[1]

    if '-' in name:
        name = name.split('-')[0]
        name = ''.join(name)
        # print(name)
    
    #for each column in sheet2
    for column in columns:
        for cell in ws2[column]:
            if cell.value == None: continue
            # if name in cell
            if name in cell.value:
                # print(name, cell.row)
                #exchange current value for id
                column_i = columns.index(column) + 1
                ws2.cell(column=column_i, row=cell.row, value = id)
                print(name)
wb.save(FILE)

#STEP 2
# copy file web_pics
# for row in id.xlsx
    #if '-' in name, remove
    # for each column
        #for each cell in column
        #if name in cell:
            #record there the id overwriting the url



#STEP 1
#launch driver
#login into wp
#go to media url
#for item in media
#   retrieve name, id
#record to excel

# OUTPUT_FILE = r'C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\login and launcher\id.xlsx'
# wb = load_workbook(filename = OUTPUT_FILE)
# ws = wb.active

# options = uc.ChromeOptions()
# # d = uc.Chrome(executable_path=EXECUTABLE_PATH,options=options)
# d = uc.Chrome(options=options)

# trs = d.find_elements(By.XPATH, '//tbody/tr')

# for tr in trs:
#     id = tr.get_attribute('id').replace('post-','')
#     text = tr.text
#     name = text.split('\n')[1]
#     print(name)
#     ws.append([name, id])
# wb.save(OUTPUT_FILE)

